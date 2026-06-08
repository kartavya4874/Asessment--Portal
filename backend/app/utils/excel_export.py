from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4A00E0", end_color="4A00E0", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
import re

def _year_sort_key(student_data: dict) -> tuple:
    """Natural sort key for year/semester fields like '1st Year', 'Sem 2', '3rd Semester', etc."""
    year = str(student_data.get("year", "")).strip()
    # Extract the first number from the string for natural ordering
    match = re.search(r'(\d+)', year)
    return (int(match.group(1)) if match else 999, year.lower())

COLUMNS = [
    ("Roll No", 15),
    ("Name", 25),
    ("Email", 35),
    ("Specialization", 20),
    ("Year/Sem", 12),
    ("Submission Files", 40),
    ("Text Answer", 40),
    ("URLs Submitted", 40),
    ("Marks", 10),
    ("Feedback", 30),
]


def _style_header(ws):
    for col_idx, (col_name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width


def _add_student_row(ws, row_idx: int, student_data: dict):
    submission = student_data.get("submission") or {}
    files = submission.get("files", [])
    file_urls = "\n".join([f"{f['name']}: {f['url']}" for f in files]) if files else "—"
    urls = "\n".join(submission.get("urls", [])) or "—"

    values = [
        student_data.get("rollNumber", ""),
        student_data.get("name", ""),
        student_data.get("email", ""),
        student_data.get("specialization", ""),
        student_data.get("year", ""),
        file_urls,
        submission.get("textAnswer", "") or "—",
        urls,
        submission.get("marks", "—") if submission.get("marks") is not None else "—",
        submission.get("feedback", "") or "—",
    ]

    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = CELL_ALIGNMENT
        cell.border = THIN_BORDER


def generate_assessment_excel(
    assessment_title: str,
    program_name: str,
    students_data: list,
) -> BytesIO:
    """Generate an Excel file for a single assessment."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{program_name[:25]} - Assessment"

    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{assessment_title} — {program_name}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="4A00E0")
    title_cell.alignment = Alignment(horizontal="center")

    # Blank row
    ws.append([])

    # Header on row 3
    _style_header_at_row(ws, 3)

    # Data rows (sorted by year/semester)
    sorted_students = sorted(students_data, key=_year_sort_key)
    for idx, student in enumerate(sorted_students):
        _add_student_row(ws, idx + 4, student)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _style_header_at_row(ws, row: int):
    for col_idx, (col_name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width


def generate_combined_excel(
    programs_data: list,
) -> BytesIO:
    """Generate a combined Excel workbook with separate sheets per program."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for program_info in programs_data:
        program_name = program_info["programName"][:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=program_name)

        _style_header(ws)

        row_idx = 2
        for assessment_group in program_info.get("assessments", []):
            # Assessment title separator
            ws.merge_cells(
                start_row=row_idx,
                start_column=1,
                end_row=row_idx,
                end_column=10,
            )
            title_cell = ws.cell(row=row_idx, column=1)
            title_cell.value = f"📝 {assessment_group['assessmentTitle']}"
            title_cell.font = Font(name="Calibri", bold=True, size=12, color="6C5CE7")
            row_idx += 1

            sorted_students = sorted(assessment_group.get("students", []), key=_year_sort_key)
            for student in sorted_students:
                _add_student_row(ws, row_idx, student)
                row_idx += 1

            row_idx += 1  # Blank row between assessments

    if not wb.sheetnames:
        wb.create_sheet(title="No Data")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─── Attendance Excel Exports ──────────────────────────────

ATTENDANCE_HEADER_FILL = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
PRESENT_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
LATE_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
ABSENT_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

ATTENDANCE_COLUMNS = [
    ("S.No", 8),
    ("Roll No", 16),
    ("Name", 28),
    ("Email", 35),
    ("Status", 12),
    ("Marked At", 22),
]


def _style_attendance_header(ws, row=1, columns=None):
    cols = columns or ATTENDANCE_COLUMNS
    for col_idx, (col_name, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = ATTENDANCE_HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width


def _get_status_fill(status):
    if status == "present":
        return PRESENT_FILL
    elif status == "late":
        return LATE_FILL
    return ABSENT_FILL


def generate_attendance_session_excel(
    session_title: str,
    session_date: str,
    slot_label: str,
    records: list,
    absent_students: list,
) -> BytesIO:
    """Generate an Excel file for a single attendance session."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Title rows
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{session_title}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1E88E5")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    subtitle = ws.cell(row=2, column=1)
    subtitle.value = f"Date: {session_date}  |  Slot: {slot_label}  |  Total Records: {len(records)}  |  Absent: {len(absent_students)}"
    subtitle.font = Font(name="Calibri", size=11, color="666666")
    subtitle.alignment = Alignment(horizontal="center")

    ws.append([])  # Blank row

    # Header on row 4
    _style_attendance_header(ws, row=4)

    # Present + Late records
    row_idx = 5
    for record in sorted(records, key=lambda r: str(r.get("rollNumber") or "")):
        status = record.get("status", "present")
        marked_at = record.get("markedAt", "")
        if marked_at:
            if hasattr(marked_at, "strftime"):
                marked_at = marked_at.strftime("%I:%M:%S %p")
            elif isinstance(marked_at, str):
                if "T" in marked_at:
                    try:
                        from datetime import datetime as dt
                        t = dt.fromisoformat(marked_at.replace("Z", "+00:00"))
                        marked_at = t.strftime("%I:%M:%S %p")
                    except Exception:
                        pass
                else:
                    marked_at = str(marked_at)
            else:
                marked_at = str(marked_at)

        values = [
            serial,
            record.get("rollNumber", ""),
            record.get("studentName", ""),
            record.get("email", ""),
            "✅ Present" if status == "present" else "⏰ Late",
            marked_at or "—",
        ]
        fill = _get_status_fill(status)
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER
            if col_idx == 5:  # Status column
                cell.fill = fill
        row_idx += 1
        serial += 1

    # Absent students
    if absent_students:
        row_idx += 1  # Blank separator
        for student in sorted(absent_students, key=lambda s: str(s.get("rollNumber") or "")):
            values = [
                serial,
                student.get("rollNumber", ""),
                student.get("studentName", student.get("name", "")),
                student.get("email", ""),
                "❌ Absent",
                "—",
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
                if col_idx == 5:
                    cell.fill = ABSENT_FILL
            row_idx += 1
            serial += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_attendance_combined_excel(
    sessions_data: list,
    all_students: list,
) -> BytesIO:
    """
    Generate a combined attendance Excel workbook with:
    - Sheet 1: Master Overview (students × sessions matrix with P/L/A)
    - Sheet per session: detailed attendance
    """
    wb = Workbook()

    # ─── Sheet 1: Master Overview ───────────────────────────
    ws_master = wb.active
    ws_master.title = "Overview"

    # Build header: S.No, Roll No, Name, [Session1, Session2, ...], Total Present, Total Late, %, Status
    session_titles = [s["title"] for s in sessions_data]
    overview_cols = [
        ("S.No", 8), ("Roll No", 16), ("Name", 25),
    ]
    for title in session_titles:
        overview_cols.append((title[:15], 14))
    overview_cols.extend([
        ("Present", 10), ("Late", 8), ("Absent", 8), ("Total", 8), ("Attendance %", 14),
    ])

    # Style header
    for col_idx, (col_name, width) in enumerate(overview_cols, 1):
        cell = ws_master.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = ATTENDANCE_HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws_master.column_dimensions[cell.column_letter].width = width

    # Build a lookup: { session_id: { student_id: status } }
    session_records_map = {}
    for s in sessions_data:
        sid = s["sessionId"]
        record_map = {}
        for r in s.get("records", []):
            record_map[r.get("studentId", "")] = r.get("status", "absent")
        session_records_map[sid] = record_map

    # One row per student
    sorted_students = sorted(all_students, key=lambda s: str(s.get("rollNumber") or ""))
    for row_offset, student in enumerate(sorted_students):
        row = row_offset + 2
        st_id = student.get("id", student.get("studentId", ""))

        ws_master.cell(row=row, column=1, value=row_offset + 1).border = THIN_BORDER
        ws_master.cell(row=row, column=2, value=student.get("rollNumber", "")).border = THIN_BORDER
        ws_master.cell(row=row, column=3, value=student.get("name", "")).border = THIN_BORDER

        present_count = 0
        late_count = 0
        absent_count = 0

        for s_idx, s in enumerate(sessions_data):
            col = 4 + s_idx
            status = session_records_map.get(s["sessionId"], {}).get(st_id, "absent")
            label = "P" if status == "present" else "L" if status == "late" else "A"

            cell = ws_master.cell(row=row, column=col, value=label)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.fill = _get_status_fill(status)

            if status == "present":
                present_count += 1
            elif status == "late":
                late_count += 1
            else:
                absent_count += 1

        total_sessions = len(sessions_data)
        attended = present_count + late_count
        pct = round((attended / total_sessions * 100), 1) if total_sessions > 0 else 0

        base_col = 4 + len(sessions_data)
        for ci, val in enumerate([present_count, late_count, absent_count, total_sessions, f"{pct}%"]):
            cell = ws_master.cell(row=row, column=base_col + ci, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # Color the percentage cell
        pct_cell = ws_master.cell(row=row, column=base_col + 4)
        if pct >= 75:
            pct_cell.fill = PRESENT_FILL
        elif pct >= 50:
            pct_cell.fill = LATE_FILL
        else:
            pct_cell.fill = ABSENT_FILL

    # ─── Per-Session Sheets ─────────────────────────────────
    import re
    for s_idx, s in enumerate(sessions_data):
        raw_name = f"{s_idx + 1}. {s['title']}"
        sheet_name = re.sub(r'[\\/*?:\[\]]', '', raw_name)[:31]
        ws = wb.create_sheet(title=sheet_name)
        _style_attendance_header(ws, row=1)

        row_idx = 2
        serial = 1
        session_map = session_records_map.get(s["sessionId"], {})

        for student in sorted_students:
            st_id = student.get("id", student.get("studentId", ""))
            status = session_map.get(st_id, "absent")

            # Find mark time from records
            marked_at = "—"
            for r in s.get("records", []):
                if r.get("studentId") == st_id:
                    raw = r.get("markedAt", "")
                    if raw:
                        if hasattr(raw, "strftime"):
                            marked_at = raw.strftime("%I:%M:%S %p")
                        elif isinstance(raw, str) and "T" in raw:
                            try:
                                from datetime import datetime as dt
                                t = dt.fromisoformat(raw.replace("Z", "+00:00"))
                                marked_at = t.strftime("%I:%M:%S %p")
                            except Exception:
                                marked_at = raw
                        else:
                            marked_at = str(raw)
                    break

            status_label = "✅ Present" if status == "present" else "⏰ Late" if status == "late" else "❌ Absent"
            values = [serial, student.get("rollNumber", ""), student.get("name", ""), student.get("email", ""), status_label, marked_at]
            fill = _get_status_fill(status)
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
                if col_idx == 5:
                    cell.fill = fill
            row_idx += 1
            serial += 1

    if not wb.sheetnames:
        wb.create_sheet(title="No Data")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
